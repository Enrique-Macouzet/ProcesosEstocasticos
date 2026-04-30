"""
modulos/importar_excel.py
Permite cargar un archivo Excel con la estructura del modelo MDP
(igual a la generada por el módulo de exportación) y cargar los datos en la sesión.
"""

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from guardado.sesion import get_mdp, reset_mdp, set_costo, set_transicion, mdp_completo

st.set_page_config(page_title="Importar desde Excel — MDP", page_icon="📥")

# ---------- ESTILOS ----------
st.markdown("""
<style>
.policy-box { background: #111827; border: 1px solid #1E2A3A; border-radius: 12px; padding: 1rem; margin-bottom: 0.8rem; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div style="margin-bottom:1.5rem;">
    <div style="font-family:'IBM Plex Mono',monospace;font-size:.72rem;color:#F5A800;letter-spacing:.15em;margin-bottom:.4rem;">MÓDULO 10</div>
    <h1 style="font-family:'Sora',sans-serif;font-weight:700;font-size:1.8rem;color:#E8EAF0;margin:0;">Importar desde Excel</h1>
    <p style="color:#8FA0B8;font-size:.9rem;margin:.4rem 0 0 0;">Carga un modelo MDP previamente guardado o creado con la plantilla.</p>
</div>
""", unsafe_allow_html=True)

# --- Descargar plantilla ---
with open("plantilla_mdp.xlsx", "rb") as f:
    st.download_button(
        label="📄 Descargar plantilla de ejemplo",
        data=f,
        file_name="plantilla_mdp.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

st.markdown("---")

archivo = st.file_uploader("Selecciona un archivo Excel (.xlsx)", type=["xlsx"])

if archivo is not None:
    if st.button("Cargar modelo desde Excel", type="primary"):
        try:
            wb = load_workbook(archivo, read_only=True)
            if "Ingreso de Datos" not in wb.sheetnames:
                st.error("El archivo no contiene la hoja 'Ingreso de Datos'. Asegúrate de usar el formato correcto.")
                st.stop()

            ws = wb["Ingreso de Datos"]

            # Leer primera fila: Estados | Decisiones | Tipo de modelo
            row1 = [cell.value for cell in ws[1]]
            estados_str = row1[0] if row1[0] else ""
            decisiones_str = row1[1] if len(row1) > 1 and row1[1] else ""
            tipo = row1[2] if len(row1) > 2 and row1[2] else "costos"

            # Validar tipo
            if tipo not in ("costos", "ganancias"):
                st.error(f"Tipo de modelo no válido: '{tipo}'. Debe ser 'costos' o 'ganancias'.")
                st.stop()

            # Limpiar modelo actual y cargar datos básicos
            reset_mdp()
            mdp = get_mdp()
            mdp["estados"] = [e.strip() for e in estados_str.split(",") if e.strip()]
            mdp["decisiones"] = [d.strip() for d in decisiones_str.split(",") if d.strip()]
            mdp["tipo"] = tipo

            # Inicializar decisiones_data vacío
            for d in mdp["decisiones"]:
                mdp["decisiones_data"][d] = {
                    "estados_afectados": [],
                    "costos": {},
                    "transiciones": {}
                }

            # Buscar la tabla de costos (empieza después de fila 1 y un espacio)
            row_idx = 2
            while row_idx <= ws.max_row and ws.cell(row=row_idx, column=1).value is None:
                row_idx += 1
            if row_idx > ws.max_row:
                st.error("No se encontró la tabla de costos.")
                st.stop()

            # Leer tabla de costos
            # Encabezados: Estado | d1 | d2 | ...
            headers = [ws.cell(row=row_idx, column=c).value for c in range(1, len(mdp["decisiones"])+2)]
            if headers[0] != "Estado":
                st.error(f"Se esperaba 'Estado' en la primera columna de la tabla de costos, pero se encontró '{headers[0]}'.")
                st.stop()
            decisiones_archivo = headers[1:]
            if decisiones_archivo != mdp["decisiones"]:
                st.error("Las decisiones en la tabla de costos no coinciden con la lista de decisiones.")
                st.stop()

            row_idx += 1
            while row_idx <= ws.max_row:
                estado = ws.cell(row=row_idx, column=1).value
                if estado is None or estado not in mdp["estados"]:
                    break   # Fin de la tabla de costos
                for j, d in enumerate(mdp["decisiones"]):
                    costo = ws.cell(row=row_idx, column=j+2).value
                    if costo is not None and costo != "—":
                        try:
                            set_costo(estado, d, float(costo))
                            # Registrar estado afectado
                            if estado not in mdp["decisiones_data"][d]["estados_afectados"]:
                                mdp["decisiones_data"][d]["estados_afectados"].append(estado)
                        except ValueError:
                            st.error(f"Valor de costo inválido en estado {estado}, decisión {d}: {costo}")
                            st.stop()
                row_idx += 1

            # Buscar matrices de transición
            while row_idx <= ws.max_row:
                celda = ws.cell(row=row_idx, column=1).value
                if celda and celda.startswith("Decisión:"):
                    decision = celda.split(":")[1].strip()
                    if decision not in mdp["decisiones"]:
                        st.error(f"Decisión '{decision}' no encontrada en la lista de decisiones.")
                        st.stop()
                    # Siguiente fila: encabezados de estados destino
                    row_idx += 1
                    dest_headers = [ws.cell(row=row_idx, column=c).value for c in range(1, len(mdp["estados"])+2)]
                    if dest_headers[0] != "Estado" or dest_headers[1:] != mdp["estados"]:
                        st.error(f"Los estados destino en la matriz de {decision} no coinciden.")
                        st.stop()
                    row_idx += 1
                    # Filas de transiciones
                    while row_idx <= ws.max_row:
                        origen = ws.cell(row=row_idx, column=1).value
                        if origen is None or origen not in mdp["estados"]:
                            break
                        for j, dest in enumerate(mdp["estados"]):
                            prob = ws.cell(row=row_idx, column=j+2).value
                            if prob is not None:
                                try:
                                    set_transicion(origen, decision, dest, float(prob))
                                    # Registrar estado afectado
                                    if origen not in mdp["decisiones_data"][decision]["estados_afectados"]:
                                        mdp["decisiones_data"][decision]["estados_afectados"].append(origen)
                                except ValueError:
                                    st.error(f"Probabilidad inválida en {origen} -> {dest} ({decision}): {prob}")
                                    st.stop()
                        row_idx += 1
                else:
                    row_idx += 1

            wb.close()
            st.success("✅ Modelo cargado correctamente desde Excel. Ya puedes usar los demás módulos.")
            if mdp_completo():
                st.info("El modelo está completo y listo para resolver.")
            else:
                st.warning("El modelo se cargó pero no está completo. Revisa los datos en Ingreso de Datos.")
        except Exception as e:
            st.error(f"Error al leer el archivo: {str(e)}")
