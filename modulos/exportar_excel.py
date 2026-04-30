"""
modulos/exportar_excel.py
Permite seleccionar las hojas a incluir y descargar un archivo Excel
con toda la información del modelo y los resultados de los algoritmos.
"""

import streamlit as st
import pandas as pd
import numpy as np
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from guardado.sesion import get_mdp
from algoritmos.exhaustiva import enumeracion_exhaustiva, generar_politicas, evaluar_politica
from algoritmos.programacion_lineal import resolver_pl
from algoritmos.mejoramiento_politicas import mejoramiento_politicas
from algoritmos.mejoramiento_politicas_descuento import mejoramiento_politicas_descuento
from algoritmos.aproximaciones_sucesivas import aproximaciones_sucesivas

st.set_page_config(page_title="Exportar a Excel — MDP", page_icon="📤")

mdp = get_mdp()

st.markdown("""
<style>
.policy-box { background: #111827; border: 1px solid #1E2A3A; border-radius: 12px; padding: 1rem; margin-bottom: 0.8rem; }
.policy-optimal { border-left: 4px solid #F5A800; background: #0f1a24; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div style="margin-bottom:1.5rem;">
    <div style="font-family:'IBM Plex Mono',monospace;font-size:.72rem;color:#F5A800;letter-spacing:.15em;margin-bottom:.4rem;">MÓDULO 09</div>
    <h1 style="font-family:'Sora',sans-serif;font-weight:700;font-size:1.8rem;color:#E8EAF0;margin:0;">Exportar a Excel</h1>
    <p style="color:#8FA0B8;font-size:.9rem;margin:.4rem 0 0 0;">Descarga todos los datos y resultados en un archivo Excel.</p>
</div>
""", unsafe_allow_html=True)

if not mdp["estados"] or not mdp["decisiones"]:
    st.warning("El modelo no tiene datos. Ve al módulo de Ingreso de Datos para comenzar.")
    st.stop()

estados = mdp["estados"]
decisiones_data = mdp["decisiones_data"]
tipo = mdp["tipo"]

# Recuperar parámetros de sesión (valores por defecto si no se han usado)
eps_as = st.session_state.get("eps_as", 0.01)
max_iter_as = st.session_state.get("max_iter_as", 100)
alpha_as = st.session_state.get("alpha_as", 1.0)
alpha_descuento = st.session_state.get("alpha_descuento", 0.9)

# --- Opciones de hojas ---
hojas_disponibles = {
    "📥 Ingreso de Datos": "ingreso",
    "🔍 Enumeración Exhaustiva": "exhaustiva",
    "📈 Programación Lineal": "programacion_lineal",
    "🔄 Mejoramiento de Políticas": "mejoramiento",
    "💲 Mej. Políticas c/ Desc.": "mejoramiento_descuento",
    "🔁 Aproximaciones Sucesivas": "aproximaciones",
    "⚖️ Comparación de Métodos": "comparacion",
}

st.markdown("### Selecciona las hojas a incluir")
seleccionadas = st.multiselect(
    "Hojas",
    options=list(hojas_disponibles.keys()),
    default=list(hojas_disponibles.keys())[:6]  # las primeras 6 por defecto
)

if st.button("📥 Generar y descargar Excel", type="primary"):
    if not seleccionadas:
        st.warning("Selecciona al menos una hoja.")
        st.stop()

    with st.spinner("Recopilando información y generando Excel..."):
        wb = Workbook()
        # Eliminar hoja por defecto
        wb.remove(wb.active)

        # ---------- HOJAS ESENCIALES ----------
        if "📥 Ingreso de Datos" in seleccionadas:
            ws = wb.create_sheet("Ingreso de Datos")
            ws.append(["Estados", "Decisiones", "Tipo de modelo"])
            ws.append([", ".join(estados), ", ".join(mdp["decisiones"]), tipo])
            ws.append([])
            ws.append(["Costos / Ganancias por estado y decisión"])
            # Tabla de costos
            headers = ["Estado"] + mdp["decisiones"]
            ws.append(headers)
            for s in estados:
                row = [s]
                for d in mdp["decisiones"]:
                    data = decisiones_data.get(d, {})
                    if s in data.get("estados_afectados", []):
                        row.append(data["costos"].get(s, 0.0))
                    else:
                        row.append("—")
                ws.append(row)
            ws.append([])
            ws.append(["Matrices de Transición por decisión"])
            for d in mdp["decisiones"]:
                ws.append([f"Decisión: {d}"])
                sub_headers = ["Estado"] + estados
                ws.append(sub_headers)
                afectados = decisiones_data[d].get("estados_afectados", [])
                for s in afectados:
                    trans = decisiones_data[d]["transiciones"].get(s, {})
                    fila = [s] + [trans.get(s2, 0.0) for s2 in estados]
                    ws.append(fila)
                ws.append([])

        # ---------- MÓDULOS DE SOLUCIÓN ----------
        # Para evitar ejecuciones múltiples que podrían fallar, los ponemos en try separados.

        if "🔍 Enumeración Exhaustiva" in seleccionadas:
            try:
                mejor_ee, resultados_ee = enumeracion_exhaustiva(estados, decisiones_data, tipo)
                if mejor_ee and not mejor_ee.get("error"):
                    ws = wb.create_sheet("Enumeración Exhaustiva")
                    ws.append(["Política óptima", "Valor esperado"])
                    nombre_ee = None
                    politicas_todas = generar_politicas(estados, decisiones_data)
                    for i, p in enumerate(politicas_todas):
                        if p == mejor_ee["politica"]:
                            nombre_ee = f"R{i+1}"
                            break
                    ws.append([f"{nombre_ee} = ({', '.join([mejor_ee['politica'][s] for s in estados])})",
                               mejor_ee["esperado"]])
                    ws.append([])
                    ws.append(["Resultados de todas las políticas evaluadas"])
                    resultados_validos = [r for r in resultados_ee if not r.get("error")]
                    if resultados_validos:
                        tab_headers = ["Política", "Decisiones", "Costo/Ganancia esperado"]
                        ws.append(tab_headers)
                        for res in resultados_validos:
                            n = None
                            for i, p in enumerate(politicas_todas):
                                if p == res["politica"]:
                                    n = f"R{i+1}"
                                    break
                            ws.append([n,
                                       ", ".join([res["politica"][s] for s in estados]),
                                       res["esperado"]])
            except Exception:
                # Si falla, no añadimos la hoja
                pass

        if "📈 Programación Lineal" in seleccionadas:
            try:
                res_pl = resolver_pl(estados, decisiones_data, tipo)
                if res_pl["exito"]:
                    ws = wb.create_sheet("Programación Lineal")
                    ws.append(["Política óptima", "Valor óptimo"])
                    ws.append([f"({', '.join([res_pl['politica'][s] for s in estados])})",
                               res_pl["valor_optimo"]])
                    ws.append([])
                    ws.append(["Función objetivo"])
                    ws.append([res_pl["modelo"]["funcion_objetivo"]])
                    ws.append([])
                    ws.append(["Restricciones"])
                    for r in res_pl["modelo"]["restricciones_desarrollo"]:
                        ws.append([r])
                    ws.append([])
                    ws.append(["Variables Y (probabilidad conjunta)"])
                    df_y = pd.DataFrame(
                        [(s, d, res_pl["variables_y"].get((s, d), 0.0))
                         for s in estados for d in mdp["decisiones"]
                         if (s, d) in res_pl["variables_y"]],
                        columns=["Estado", "Decisión", "Y"]
                    )
                    for row in dataframe_to_rows(df_y.pivot(index="Estado", columns="Decisión", values="Y").fillna(0), index=True, header=True):
                        ws.append(row)
                    ws.append([])
                    ws.append(["Coeficientes D"])
                    for row in dataframe_to_rows(res_pl["D"].fillna(0), index=True, header=True):
                        ws.append(row)
            except Exception:
                pass

        if "🔄 Mejoramiento de Políticas" in seleccionadas:
            try:
                pol_inicial = generar_politicas(estados, decisiones_data)[0]
                iter_mp = mejoramiento_politicas(estados, decisiones_data, tipo, pol_inicial)
                ws = wb.create_sheet("Mejoramiento de Políticas")
                ultima = iter_mp[-1]
                ws.append(["Política óptima", "g (costo/ganancia promedio)"])
                ws.append([f"({', '.join([ultima['nueva_politica'][s] for s in estados])})", ultima["g"]])
                ws.append([])
                ws.append(["Iteraciones", "Política evaluada", "g", "Nueva política"])
                for it in iter_mp:
                    ws.append([it.get("iter", ""),
                               ", ".join([f"{s}→{d}" for s, d in it["politica"].items()]),
                               it["g"],
                               ", ".join([f"{s}→{d}" for s, d in it.get("nueva_politica", {}).items()])])
            except Exception:
                pass

        if "💲 Mej. Políticas c/ Desc." in seleccionadas:
            try:
                pol_inicial = generar_politicas(estados, decisiones_data)[0]
                iter_mpd = mejoramiento_politicas_descuento(estados, decisiones_data, tipo, pol_inicial, alpha_descuento)
                ws = wb.create_sheet("Mej. Políticas c/ Descuento")
                ultima = iter_mpd[-1]
                ws.append(["α", "Política óptima", "Valores V finales"])
                ws.append([alpha_descuento,
                           ", ".join([f"{s}→{d}" for s, d in ultima["nueva_politica"].items()]),
                           ", ".join([f"V({s})={ultima['V'][s]:.4f}" for s in estados])])
                ws.append([])
                ws.append(["Iteraciones", "Política evaluada", "V", "Nueva política"])
                for it in iter_mpd:
                    ws.append([it.get("iter", ""),
                               ", ".join([f"{s}→{d}" for s, d in it["politica"].items()]),
                               ", ".join([f"{v:.4f}" for v in it["V"].values()]),
                               ", ".join([f"{s}→{d}" for s, d in it.get("nueva_politica", {}).items()])])
            except Exception:
                pass

        if "🔁 Aproximaciones Sucesivas" in seleccionadas:
            try:
                iter_as = aproximaciones_sucesivas(estados, decisiones_data, tipo, eps_as, max_iter_as, alpha_as)
                ws = wb.create_sheet("Aprox. Sucesivas")
                ultima = iter_as[-1]
                ws.append(["ε", "Máx. iteraciones", "α", "Política óptima"])
                ws.append([eps_as, max_iter_as, alpha_as,
                           ", ".join([f"{s}→{d}" for s, d in ultima["paso2"]["politica"].items()])])
                ws.append([])
                ws.append(["Iteración", "V", "Política"])
                for it in iter_as:
                    ws.append([it["iter"],
                               ", ".join([f"{v:.4f}" for v in it["paso1"]["V"].values()]),
                               ", ".join([f"{s}→{d}" for s, d in it["paso2"]["politica"].items()])])
            except Exception:
                pass

        if "⚖️ Comparación de Métodos" in seleccionadas:
            # Recolectamos resultados como en comparacion_metodos.py pero resumido
            resultados_comp = {}
            detalles_comp = {}
            politicas_todas = generar_politicas(estados, decisiones_data)
            def obtener_nombre(pol):
                if pol is None: return "—"
                for i, p in enumerate(politicas_todas):
                    if p == pol: return f"R{i+1}"
                return None
            try:
                mejor_ee, _ = enumeracion_exhaustiva(estados, decisiones_data, tipo)
                if mejor_ee and not mejor_ee.get("error"):
                    resultados_comp["Enumeración Exhaustiva"] = mejor_ee["politica"]
                    detalles_comp["Enumeración Exhaustiva"] = {"costo": mejor_ee["esperado"], "iteraciones": "—"}
            except: pass
            try:
                res_pl = resolver_pl(estados, decisiones_data, tipo)
                if res_pl["exito"]:
                    resultados_comp["Programación Lineal"] = res_pl["politica"]
                    detalles_comp["Programación Lineal"] = {"costo": res_pl["valor_optimo"], "iteraciones": "—"}
            except: pass
            try:
                iter_mp = mejoramiento_politicas(estados, decisiones_data, tipo, politicas_todas[0])
                resultados_comp["Mejoramiento de Políticas"] = iter_mp[-1]["nueva_politica"]
                detalles_comp["Mejoramiento de Políticas"] = {"costo": iter_mp[-1]["g"], "iteraciones": len(iter_mp)}
            except: pass
            try:
                for pol in politicas_todas[:5]:
                    try:
                        iter_mpd = mejoramiento_politicas_descuento(estados, decisiones_data, tipo, pol, alpha_descuento)
                        resultados_comp["Mej. Políticas c/ Desc."] = iter_mpd[-1]["nueva_politica"]
                        detalles_comp["Mej. Políticas c/ Desc."] = {"costo": None, "iteraciones": len(iter_mpd), "extra": iter_mpd[-1]["V"]}
                        break
                    except np.linalg.LinAlgError: continue
            except: pass
            try:
                iter_as = aproximaciones_sucesivas(estados, decisiones_data, tipo, eps_as, max_iter_as, alpha_as)
                resultados_comp["Aproximaciones Sucesivas"] = iter_as[-1]["paso2"]["politica"]
                detalles_comp["Aproximaciones Sucesivas"] = {"costo": None, "iteraciones": len(iter_as), "extra": iter_as[-1]["paso1"]["V"]}
            except: pass

            ws = wb.create_sheet("Comparación de Métodos")
            ws.append(["Método", "Política", "Costo/Ganancia nativo", "Iteraciones"])
            for metodo in ["Enumeración Exhaustiva", "Programación Lineal", "Mejoramiento de Políticas",
                           "Mej. Políticas c/ Desc.", "Aproximaciones Sucesivas"]:
                pol = resultados_comp.get(metodo)
                det = detalles_comp.get(metodo, {})
                nombre_pol = obtener_nombre(pol) if pol else "—"
                pol_str = f"{nombre_pol} = ({', '.join([pol[s] for s in estados])})" if pol else "No disponible"
                costo = det.get("costo")
                iters = det.get("iteraciones")
                costo_str = f"{costo:.6f}" if costo is not None else "—"
                iter_str = str(iters) if iters is not None else "—"
                ws.append([metodo, pol_str, costo_str, iter_str])

        # Guardar en BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

    st.success("Excel generado correctamente.")
    st.download_button(
        label="⬇️ Descargar Excel",
        data=output,
        file_name="resultados_mdp.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
